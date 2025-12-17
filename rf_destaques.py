import streamlit as st
import pandas as pd
from datetime import datetime, date
from io import BytesIO
import re
from openpyxl import load_workbook

SHEET_NAME = "Crédito bancário"

st.set_page_config(page_title="RF | Destaques Crédito Bancário", layout="wide")
st.title("RF | Destaques Crédito Bancário")
st.caption(
    'Lê apenas a aba "Crédito bancário" (cabeçalho fixo na linha 6) e monta Top 5 por indexador e prazo. '
    'Exibe taxas, aplicação mínima e vencimento formatados e gera mensagem pronta para WhatsApp.'
)

# =============================
# Helpers
# =============================
def normalize_colname(c):
    if c is None:
        return ""
    return str(c).strip().replace("\n", " ").replace("\r", " ")

def find_col(df, candidates):
    for cand in candidates:
        cand_l = cand.lower()
        for c in df.columns:
            if cand_l == c.lower() or cand_l in c.lower():
                return c
    return None

def to_numeric_series(s):
    if s is None:
        return pd.Series(dtype="float64")
    if pd.api.types.is_numeric_dtype(s):
        return pd.to_numeric(s, errors="coerce")

    s = s.astype(str).str.strip()
    s = s.str.replace(".", "", regex=False)
    s = s.str.replace(",", ".", regex=False)
    s = s.str.extract(r"(-?\d+(\.\d+)?)", expand=True)[0]
    return pd.to_numeric(s, errors="coerce")

def to_date_series(s):
    if s is None:
        return pd.Series(dtype="datetime64[ns]")
    return pd.to_datetime(s, errors="coerce", dayfirst=True)

def categorize_horizon(days):
    if pd.isna(days):
        return None
    if days <= 360:
        return "Curto (até 360d)"
    if days <= 1080:
        return "Médio (361 a 1080d)"
    return "Longo (acima de 1080d)"

def classify_indexer(raw):
    if raw is None or pd.isna(raw):
        return None
    t = str(raw).upper()
    if "IPCA" in t:
        return "IPCA"
    if "CDI" in t or "PÓS" in t or "POS" in t:
        return "Pós (CDI)"
    if "PRÉ" in t or "PRE" in t or "FIXA" in t:
        return "Pré"
    return None

def parse_rate_value(x):
    if x is None or pd.isna(x):
        return None
    if isinstance(x, (int, float)):
        return float(x)

    s = str(x).upper().replace("%", "").replace(" ", "")
    m = re.search(r"(-?\d[\d\.,]*)", s)
    if not m:
        return None
    num = m.group(1)
    if "." in num and "," in num:
        num = num.replace(".", "").replace(",", ".")
    elif "," in num:
        num = num.replace(",", ".")
    return float(num)

def format_rate_for_display(rate_num, indexador):
    if rate_num is None or pd.isna(rate_num):
        return ""
    val = float(rate_num)

    if indexador == "Pós (CDI)":
        val = val * 100 if val <= 2 else val
        return f"{val:,.2f}%".replace(",", "X").replace(".", ",").replace("X", ".")

    if val <= 1.5:
        val = val * 100
    return f"{val:.2f}%".replace(".", ",")

def format_currency_brl(value):
    if value is None or pd.isna(value):
        return ""
    try:
        v = float(value)
    except:
        return ""
    return f"R$ {int(round(v)):,.0f}".replace(",", ".")

def format_date_br(dt):
    if dt is None or pd.isna(dt):
        return ""
    return pd.to_datetime(dt).strftime("%d/%m/%Y")

def top_n_block(df, idx, horizon, n):
    sub = df[(df["indexador_pad"] == idx) & (df["horizonte"] == horizon)]
    return sub.sort_values("taxa_num", ascending=False).head(int(n))

def download_csv(df):
    return df.to_csv(index=False).encode("utf-8-sig")

# =============================
# Fast Excel reader
# =============================
@st.cache_data(show_spinner=False)
def read_credito_bancario_fast(file_bytes):
    bio = BytesIO(file_bytes)
    wb = load_workbook(bio, read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f'Não encontrei a aba "{SHEET_NAME}". Abas disponíveis: {wb.sheetnames}')

    ws = wb[SHEET_NAME]

    HEADER_ROW = 6
    header = [normalize_colname(c.value) for c in ws[HEADER_ROW]]

    data = []
    empty_streak = 0
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        if row is None or all(x is None or str(x).strip() == "" for x in row):
            empty_streak += 1
            if empty_streak >= 20:
                break
            continue
        empty_streak = 0
        data.append(row)

    df = pd.DataFrame(data, columns=header)
    return df.dropna(axis=1, how="all")

# =============================
# UI
# =============================
uploaded = st.file_uploader("Envie a planilha (.xlsx ou .xlsm)", type=["xlsx", "xlsm"])

with st.sidebar:
    st.header("Configurações")
    top_n = st.number_input("Top N por bloco", min_value=1, max_value=20, value=5, step=1)

    st.subheader("Filtros opcionais")
    use_rating_filter = st.checkbox("Aplicar rating mínimo", value=False)
    rating_min = st.text_input("Rating mínimo (ex: A-, A, AA-)", value="A-")

    use_min_app_filter = st.checkbox("Filtrar por aplicação mínima máxima", value=False)
    max_min_app = st.number_input("Aplicação mínima máxima (R$)", min_value=0, value=0, step=1000)

if not uploaded:
    st.info("Envie o arquivo para começar.")
    st.stop()

df = read_credito_bancario_fast(uploaded.getvalue())

# Detect columns
col_emissor = find_col(df, ["Emissor"])
col_produto = find_col(df, ["Produto"])
col_indexador = find_col(df, ["Indexador"])
col_taxa = find_col(df, ["Tx", "Taxa", "Máxima", "Maxima"])
col_prazo = find_col(df, ["Prazo"])
col_venc = find_col(df, ["Vencimento"])
col_min = find_col(df, ["Aplicação", "Aplicacao", "mínima", "minima"])
col_rating = find_col(df, ["Rating"])

missing = []
if not col_emissor:
    missing.append("Emissor")
if not col_produto:
    missing.append("Produto")
if not col_indexador:
    missing.append("Indexador")
if not col_taxa:
    missing.append("Tx. Máxima/Taxa")
if not col_prazo and not col_venc:
    missing.append("Prazo ou Vencimento")
if not col_min:
    missing.append("Aplicação mínima")
if not col_venc:
    missing.append("Vencimento")

if missing:
    st.error("Colunas obrigatórias não encontradas: " + ", ".join(missing))
    st.write("Colunas detectadas:", list(df.columns))
    st.stop()

# =============================
# Transform
# =============================
df["indexador_pad"] = df[col_indexador].apply(classify_indexer)

if col_prazo:
    df["prazo_dias"] = to_numeric_series(df[col_prazo])
else:
    venc_dt = to_date_series(df[col_venc])
    df["prazo_dias"] = (venc_dt - pd.Timestamp(date.today())).dt.days

df["horizonte"] = df["prazo_dias"].apply(categorize_horizon)

df["taxa_num"] = df[col_taxa].apply(parse_rate_value)
df["taxa_fmt"] = df.apply(lambda r: format_rate_for_display(r["taxa_num"], r["indexador_pad"]), axis=1)

df["aplic_min_num"] = to_numeric_series(df[col_min])
df["aplic_min_fmt"] = df["aplic_min_num"].apply(format_currency_brl)

df["_venc_dt"] = to_date_series(df[col_venc])
df["venc_fmt"] = df["_venc_dt"].apply(format_date_br)

# filtros
df = df[df["indexador_pad"].notna() & df["horizonte"].notna() & df["taxa_num"].notna()].copy()

# rating filter
if use_rating_filter and col_rating:
    rating_map = {
        "AAA": 1, "AA+": 2, "AA": 3, "AA-": 4,
        "A+": 5, "A": 6, "A-": 7,
        "BBB+": 8, "BBB": 9, "BBB-": 10,
        "BB+": 11, "BB": 12, "BB-": 13,
        "B+": 14, "B": 15, "B-": 16,
        "CCC": 17, "CC": 18, "C": 19, "D": 20
    }

    def rating_score(x):
        if x is None or pd.isna(x):
            return None
        t = str(x).strip().upper().replace(" ", "")
        return rating_map.get(t)

    df["_rating_score"] = df[col_rating].apply(rating_score)
    min_score = rating_score(rating_min)
    if min_score is not None:
        df = df[df["_rating_score"].notna() & (df["_rating_score"] <= min_score)].copy()

# min app filter
if use_min_app_filter and max_min_app and max_min_app > 0:
    df = df[df["aplic_min_num"].notna() & (df["aplic_min_num"] <= float(max_min_app))].copy()

# =============================
# Display
# =============================
st.subheader("Base tratada (preview)")
preview_cols = [col_emissor, col_produto, col_indexador, "taxa_fmt", "aplic_min_fmt", "venc_fmt"]
if col_rating:
    preview_cols.append(col_rating)
preview_cols += ["prazo_dias", "horizonte"]
st.dataframe(df[preview_cols].head(80), use_container_width=True, height=340)

indexers = ["Pós (CDI)", "Pré", "IPCA"]
horizons = ["Curto (até 360d)", "Médio (361 a 1080d)", "Longo (acima de 1080d)"]

st.divider()
st.subheader("Top do dia (Top N por indexador e horizonte)")

tabs = st.tabs(indexers)
for i, idx in enumerate(indexers):
    with tabs[i]:
        cols = st.columns(3)
        for j, hz in enumerate(horizons):
            with cols[j]:
                st.markdown(f"### {hz}")
                b = top_n_block(df, idx, hz, top_n)
                if b.empty:
                    st.info("Sem ativos")
                else:
                    show_cols = [col_emissor, col_produto, col_indexador, "taxa_fmt", "aplic_min_fmt", "venc_fmt"]
                    if col_rating:
                        show_cols.append(col_rating)
                    st.dataframe(b[show_cols], use_container_width=True, height=260)

# =============================
# WhatsApp message (Top N por indexador)
# =============================
st.divider()
st.subheader("Mensagem pronta para WhatsApp (Top N por indexador)")

def build_whatsapp_message(df, top_n):
    data_envio = datetime.now().strftime("%d/%m/%Y")

    def format_card(row, prefixo_taxa=""):
        emissor = str(row.get(col_emissor, "")).strip()
        produto = str(row.get(col_produto, "")).strip()
        taxa = str(row.get("taxa_fmt", "")).strip()
        venc = str(row.get("venc_fmt", "")).strip()
        amin = str(row.get("aplic_min_fmt", "")).strip()

        titulo = f"{produto} {emissor}".strip()
        taxa_exibicao = f"{prefixo_taxa}{taxa}" if taxa else ""

        return (
            f"🏦*{titulo}*\n"
            f"⏰ Vencimento: {venc}\n"
            f"📈 Taxa: {taxa_exibicao}\n"
            f"💰mínimo: {amin}\n"
        )

    def section(indexador_label, section_title, prefixo_taxa=""):
        sub = df[df["indexador_pad"] == indexador_label].copy()
        sub = sub.sort_values("taxa_num", ascending=False).head(int(top_n))

        if sub.empty:
            return f"📍*{section_title}*\n- (sem ativos hoje)\n\n"

        cards = "\n".join(format_card(r, prefixo_taxa=prefixo_taxa) for _, r in sub.iterrows())
        return f"📍*{section_title}*\n{cards}\n"

    msg = (
        "*Destaques de ativos Bancários*\n"
        f"🚨*TAXAS DE HOJE ({data_envio})*\n\n"
        + section("Pós (CDI)", "PÓS-FIXADOS")
        + section("Pré", "PRÉ-FIXADOS")
        + section("IPCA", "IPCA", prefixo_taxa="IPCA+ ")
    )

    return msg

msg = build_whatsapp_message(df, top_n=top_n)

st.text_area("Copie e cole no WhatsApp", value=msg, height=560)

st.download_button(
    "Baixar mensagem (.txt)",
    data=msg.encode("utf-8"),
    file_name=f"destaques_bancarios_{datetime.now().strftime('%Y%m%d')}.txt",
    mime="text/plain",
)

# =============================
# Download consolidado
# =============================
st.divider()
st.subheader("Download do consolidado (Top N por bloco)")

blocks = []
for idx in indexers:
    for hz in horizons:
        b = top_n_block(df, idx, hz, top_n).copy()
        if not b.empty:
            b["Bloco"] = f"{idx} | {hz}"
            blocks.append(b)

result = pd.concat(blocks, ignore_index=True) if blocks else pd.DataFrame()

if result.empty:
    st.info("Não há dados suficientes para gerar o consolidado.")
else:
    st.download_button(
        "Baixar CSV consolidado",
        data=download_csv(result),
        file_name=f"top_ativos_credito_bancario_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
        mime="text/csv",
    )
