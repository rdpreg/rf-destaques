import streamlit as st
import pandas as pd
from datetime import datetime, date
from io import BytesIO
import re
from openpyxl import load_workbook
import streamlit.components.v1 as components

SHEET_NAME = "Crédito bancário"
SHEET_PUBLICOS = "Títulos Públicos"

st.set_page_config(page_title="RF | Destaques (V2)", layout="wide")
st.title("RF | Destaques RF (V2)")
st.caption("Crédito bancário + Títulos públicos (mensagens prontas para WhatsApp)")

# =============================
# Helpers
# =============================
def normalize_colname(c):
    if c is None:
        return ""
    return str(c).strip().replace("\n", " ").replace("\r", " ")

def find_col(df, candidates):
    for cand in candidates:
        cand_l = cand.lower()
        for c in df.columns:
            c_l = str(c).lower()
            if cand_l == c_l or cand_l in c_l:
                return c
    return None

def to_numeric_series(s):
    if s is None:
        return pd.Series(dtype="float64")
    if pd.api.types.is_numeric_dtype(s):
        return pd.to_numeric(s, errors="coerce")
    s = s.astype(str).str.strip()
    s = s.str.replace(".", "", regex=False)
    s = s.str.replace(",", ".", regex=False)
    s = s.str.extract(r"(-?\d+(\.\d+)?)", expand=True)[0]
    return pd.to_numeric(s, errors="coerce")

def to_date_series(s):
    if s is None:
        return pd.Series(dtype="datetime64[ns]")
    return pd.to_datetime(s, errors="coerce", dayfirst=True)

def categorize_horizon(days):
    if pd.isna(days):
        return None
    if days <= 360:
        return "Curto (até 360d)"
    if days <= 1080:
        return "Médio (361 a 1080d)"
    return "Longo (acima de 1080d)"

def classify_indexer_bancario(raw):
    if raw is None or pd.isna(raw):
        return None
    t = str(raw).upper()
    if "IPCA" in t:
        return "IPCA"
    if "CDI" in t or "PÓS" in t or "POS" in t:
        return "Pós (CDI)"
    if "PRÉ" in t or "PRE" in t or "FIXA" in t:
        return "Pré"
    return None

def parse_rate_value(x):
    if x is None or pd.isna(x):
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).upper().replace("%", "").replace(" ", "")
    m = re.search(r"(-?\d[\d\.,]*)", s)
    if not m:
        return None
    num = m.group(1)
    if "." in num and "," in num:
        num = num.replace(".", "").replace(",", ".")
    elif "," in num:
        num = num.replace(",", ".")
    return float(num)

def format_rate_for_display(rate_num, indexador):
    if rate_num is None or pd.isna(rate_num):
        return ""
    val = float(rate_num)

    if indexador == "Pós (CDI)":
        val = val * 100 if val <= 2 else val
        return f"{val:,.2f}%".replace(",", "X").replace(".", ",").replace("X", ".")

    if val <= 1.5:
        val = val * 100
    return f"{val:.2f}%".replace(".", ",")

def format_currency_brl(value):
    if value is None or pd.isna(value):
        return ""
    try:
        v = float(value)
    except:
        return ""
    return f"R$ {int(round(v)):,.0f}".replace(",", ".")

def format_date_br(dt):
    if dt is None or pd.isna(dt):
        return ""
    return pd.to_datetime(dt).strftime("%d/%m/%Y")

def copy_button(text: str, label: str = "Copiar"):
    safe = text.replace("\\", "\\\\").replace("`", "\\`").replace("${", "\\${")
    html = f"""
    <button onclick="navigator.clipboard.writeText(`{safe}`)"
    style="cursor:pointer;padding:8px 12px;border-radius:8px;border:1px solid #ddd;background:white;">
    📋 {label}
    </button>
    """
    components.html(html, height=45)

# =============================
# Excel reader (header fixo)
# =============================
@st.cache_data(show_spinner=False)
def read_sheet_fast(file_bytes, sheet_name: str, header_row: int):
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f'Aba "{sheet_name}" não encontrada. Abas: {wb.sheetnames}')
    ws = wb[sheet_name]

    header = [normalize_colname(c.value) for c in ws[header_row]]

    data = []
    empty_streak = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row is None or all(x is None or str(x).strip() == "" for x in row):
            empty_streak += 1
            if empty_streak >= 20:
                break
            continue
        empty_streak = 0
        data.append(row)

    df = pd.DataFrame(data, columns=header)
    return df.dropna(axis=1, how="all")

# =============================
# UI
# =============================
uploaded = st.file_uploader("Envie a planilha (.xlsx ou .xlsm)", type=["xlsx", "xlsm"])

with st.sidebar:
    st.header("Configurações")
    top_n = st.number_input("Top N por bloco (Crédito bancário)", min_value=1, max_value=20, value=5, step=1)
    mostrar_apenas_blocos_com_ativos = st.checkbox("Mostrar apenas blocos com ativos", value=True)

if not uploaded:
    st.info("Envie o arquivo para começar.")
    st.stop()

file_bytes = uploaded.getvalue()

tab_banc, tab_pub = st.tabs(["Crédito bancário", "Títulos públicos (NTN-B)"])

# =========================================================
# TAB 1: Crédito bancário (igual ao que você tinha)
# =========================================================
with tab_banc:
    st.subheader("Crédito bancário")

    df = read_sheet_fast(file_bytes, SHEET_NAME, header_row=6)

    col_emissor = find_col(df, ["Emissor"])
    col_produto = find_col(df, ["Produto"])
    col_indexador = find_col(df, ["Indexador"])
    col_taxa = find_col(df, ["Tx. Portal", "Taxa Portal"])
    col_prazo = find_col(df, ["Prazo"])
    col_venc = find_col(df, ["Vencimento"])
    col_min = find_col(df, ["Aplicação", "Aplicacao", "mínima", "minima"])

    missing = []
    if not col_emissor: missing.append("Emissor")
    if not col_produto: missing.append("Produto")
    if not col_indexador: missing.append("Indexador")
    if not col_taxa: missing.append("Tx. Portal/Taxa Portal")
    if not col_prazo and not col_venc: missing.append("Prazo ou Vencimento")
    if not col_min: missing.append("Aplicação mínima")
    if not col_venc: missing.append("Vencimento")

    if missing:
        st.error("Colunas obrigatórias não encontradas: " + ", ".join(missing))
        st.write("Colunas detectadas:", list(df.columns))
        st.stop()

    df["indexador_pad"] = df[col_indexador].apply(classify_indexer_bancario)

    if col_prazo:
        df["prazo_dias"] = to_numeric_series(df[col_prazo])
    else:
        venc_dt = to_date_series(df[col_venc])
        df["prazo_dias"] = (venc_dt - pd.Timestamp(date.today())).dt.days

    df["horizonte"] = df["prazo_dias"].apply(categorize_horizon)

    df["taxa_num"] = df[col_taxa].apply(parse_rate_value)
    df["taxa_fmt"] = df.apply(lambda r: format_rate_for_display(r["taxa_num"], r["indexador_pad"]), axis=1)

    df["aplic_min_num"] = to_numeric_series(df[col_min])
    df["aplic_min_fmt"] = df["aplic_min_num"].apply(format_currency_brl)

    df["_venc_dt"] = to_date_series(df[col_venc])
    df["venc_fmt"] = df["_venc_dt"].apply(format_date_br)

    df = df[df["indexador_pad"].notna() & df["horizonte"].notna() & df["taxa_num"].notna()].copy()

    st.markdown("##### Base tratada (preview)")
    st.dataframe(
        df[[col_emissor, col_produto, col_indexador, "taxa_fmt", "aplic_min_fmt", "venc_fmt", "horizonte"]].head(80),
        use_container_width=True,
        height=340,
    )

    def top_n_block(dff, idx, horizon, n):
        sub = dff[(dff["indexador_pad"] == idx) & (dff["horizonte"] == horizon)]
        return sub.sort_values("taxa_num", ascending=False).head(int(n))

    st.divider()
    st.subheader("Top do dia (Top N por indexador e horizonte)")

    indexers = ["Pós (CDI)", "Pré", "IPCA"]
    horizons = ["Curto (até 360d)", "Médio (361 a 1080d)", "Longo (acima de 1080d)"]

    tabs = st.tabs(indexers)
    for i, idx in enumerate(indexers):
        with tabs[i]:
            cols = st.columns(3)
            for j, hz in enumerate(horizons):
                with cols[j]:
                    st.markdown(f"### {hz}")
                    b = top_n_block(df, idx, hz, top_n)
                    if b.empty:
                        st.info("Sem ativos")
                    else:
                        st.dataframe(
                            b[[col_emissor, col_produto, col_indexador, "taxa_fmt", "aplic_min_fmt", "venc_fmt"]],
                            use_container_width=True,
                            height=260,
                        )

    # WhatsApp: 3 campos, Top 5 por prazo
    st.divider()
    st.subheader("Mensagens prontas para WhatsApp (Crédito bancário)")

    TOP_WA = 5

    def format_card(row, prefixo_taxa=""):
        emissor = str(row.get(col_emissor, "")).strip()
        produto = str(row.get(col_produto, "")).strip()
        taxa = str(row.get("taxa_fmt", "")).strip()
        venc = str(row.get("venc_fmt", "")).strip()
        amin = str(row.get("aplic_min_fmt", "")).strip()

        titulo = f"{produto} {emissor}".strip()
        taxa_exibicao = f"{prefixo_taxa}{taxa}" if taxa else ""

        return (
            f"🏦*{titulo}*\n"
            f"⏰ Vencimento: {venc}\n"
            f"📈 Taxa: {taxa_exibicao}\n"
            f"💰mínimo: {amin}\n"
        )

    def build_message_bancario(indexador_label, titulo_indexador, prefixo_taxa=""):
        data_envio = datetime.now().strftime("%d/%m/%Y")

        msg = (
            "*Destaques de ativos Bancários*\n"
            f"🚨*TAXAS DE HOJE ({data_envio})*\n\n"
            f"📍*{titulo_indexador}*\n\n"
        )

        for hz_label, hz_title in [
            ("Curto (até 360d)", "Curto Prazo (até 360d)"),
            ("Médio (361 a 1080d)", "Médio Prazo (361 a 1080d)"),
            ("Longo (acima de 1080d)", "Longo Prazo (acima de 1080d)"),
        ]:
            sub = top_n_block(df, indexador_label, hz_label, TOP_WA)
            if sub.empty and mostrar_apenas_blocos_com_ativos:
                continue

            msg += f"*{hz_title}*\n\n"
            if sub.empty:
                msg += "- (sem ativos hoje)\n\n"
            else:
                for _, r in sub.iterrows():
                    msg += format_card(r, prefixo_taxa=prefixo_taxa) + "\n"

        return msg

    msg_pos = build_message_bancario("Pós (CDI)", "PÓS-FIXADOS")
    msg_pre = build_message_bancario("Pré", "PRÉ-FIXADOS")
    msg_ipca = build_message_bancario("IPCA", "IPCA", prefixo_taxa="IPCA+ ")

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown("### Pós-fixados")
        st.text_area("Mensagem Pós", value=msg_pos, height=560)
        copy_button(msg_pos, "Copiar Pós")
    with c2:
        st.markdown("### Pré-fixados")
        st.text_area("Mensagem Pré", value=msg_pre, height=560)
        copy_button(msg_pre, "Copiar Pré")
    with c3:
        st.markdown("### IPCA")
        st.text_area("Mensagem IPCA", value=msg_ipca, height=560)
        copy_button(msg_ipca, "Copiar IPCA")

# =========================================================
# TAB 2: Títulos públicos (somente NTN-B, listar todas)
# =========================================================
with tab_pub:
    st.subheader("Títulos públicos (Cliente | somente NTN-B | listar todas)")

    # Header row da aba de títulos costuma ser 5 nesse arquivo
    dfp = read_sheet_fast(file_bytes, SHEET_PUBLICOS, header_row=5)

    col_titulo = find_col(dfp, ["Título"])
    col_venc = find_col(dfp, ["Vencimento"])
    col_taxa = find_col(dfp, ["Taxa do portal às 10h", "Taxa do portal às 10h ¹", "Taxa do portal"])

    missing_pub = []
    if not col_titulo: missing_pub.append("Título")
    if not col_venc: missing_pub.append("Vencimento")
    if not col_taxa: missing_pub.append("Taxa do portal às 10h")

    if missing_pub:
        st.error("Títulos públicos: colunas obrigatórias não encontradas: " + ", ".join(missing_pub))
        st.write("Colunas detectadas:", list(dfp.columns))
        st.stop()

    # Filtro: somente NTN-B
    dfp = dfp[dfp[col_titulo].astype(str).str.upper().str.contains("NTN-B")].copy()

    dfp["_venc_dt"] = to_date_series(dfp[col_venc])
    dfp["venc_fmt"] = dfp["_venc_dt"].apply(format_date_br)

    dfp["prazo_dias"] = (dfp["_venc_dt"] - pd.Timestamp(date.today())).dt.days
    dfp["horizonte"] = dfp["prazo_dias"].apply(categorize_horizon)

    dfp["taxa_num"] = dfp[col_taxa].apply(parse_rate_value)
    dfp["taxa_fmt"] = dfp["taxa_num"].apply(lambda x: format_rate_for_display(x, indexador="IPCA"))

    dfp = dfp[dfp["horizonte"].notna() & dfp["taxa_num"].notna()].copy()

    st.markdown("##### Preview (somente NTN-B)")
    st.dataframe(
        dfp[[col_titulo, "venc_fmt", "taxa_fmt", "horizonte", "prazo_dias"]]
        .sort_values("prazo_dias")
        .head(80),
        use_container_width=True,
        height=340,
    )

    def pub_block_all(hz):
        sub = dfp[dfp["horizonte"] == hz].copy()
        return sub.sort_values("prazo_dias")

    def format_card_pub(row):
        titulo = str(row.get(col_titulo, "")).strip()
        venc = str(row.get("venc_fmt", "")).strip()
        taxa = str(row.get("taxa_fmt", "")).strip()
        return (
            f"🏛️*{titulo}*\n"
            f"⏰ Vencimento: {venc}\n"
            f"📈 Taxa: IPCA+ {taxa}\n"
        )

    def build_message_pub_ntnb_all():
        hoje = datetime.now().strftime("%d/%m/%Y")
        msg = (
            "*Destaques de Títulos Públicos*\n"
            f"🚨*TAXAS DE HOJE ({hoje})*\n\n"
            "📍*TESOURO IPCA+ (NTN-B)*\n\n"
        )

        for hz_label, hz_title in [
            ("Curto (até 360d)", "Curto Prazo (até 360d)"),
            ("Médio (361 a 1080d)", "Médio Prazo (361 a 1080d)"),
            ("Longo (acima de 1080d)", "Longo Prazo (acima de 1080d)"),
        ]:
            sub = pub_block_all(hz_label)

            if sub.empty and mostrar_apenas_blocos_com_ativos:
                continue

            msg += f"*{hz_title}*\n\n"
            if sub.empty:
                msg += "- (sem títulos hoje)\n\n"
            else:
                for _, r in sub.iterrows():
                    msg += format_card_pub(r) + "\n"
                msg += "\n"

        return msg

    msg_pub_ntnb = build_message_pub_ntnb_all()

    st.divider()
    st.subheader("Mensagem pronta para WhatsApp – Tesouro IPCA+ (todas as NTN-B)")
    st.text_area("Mensagem Tesouro IPCA+", value=msg_pub_ntnb, height=560)
    copy_button(msg_pub_ntnb, "Copiar Tesouro IPCA+")
