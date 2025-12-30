import streamlit as st
import pandas as pd
from datetime import datetime, date
from io import BytesIO
import re
import time
import requests
from openpyxl import load_workbook
import streamlit.components.v1 as components

# =============================
# CONSTANTS
# =============================
SHEET_BANCARIO = "Crédito bancário"
HEADER_BANCARIO = 6

SHEET_PUBLICOS = "Títulos Públicos"
HEADER_PUBLICOS = 5  # ajuste se necessário no seu arquivo

# =============================
# PAGE
# =============================
st.set_page_config(page_title="RF | Destaques (V2)", layout="wide")
st.title("RF | Destaques RF (V2)")
st.caption(
    "Crédito bancário + Títulos públicos (NTN-B) com mensagens prontas para WhatsApp "
    "e envio via Z-API usando secrets."
)

# =============================
# Helpers
# =============================
def normalize_colname(c):
    if c is None:
        return ""
    return str(c).strip().replace("\n", " ").replace("\r", " ")

def find_col(df, candidates):
    for cand in candidates:
        cand_l = cand.lower()
        for c in df.columns:
            c_l = str(c).lower()
            if cand_l == c_l or cand_l in c_l:
                return c
    return None

def to_numeric_series(s):
    if s is None:
        return pd.Series(dtype="float64")
    if pd.api.types.is_numeric_dtype(s):
        return pd.to_numeric(s, errors="coerce")
    s = s.astype(str).str.strip()
    s = s.str.replace(".", "", regex=False)
    s = s.str.replace(",", ".", regex=False)
    s = s.str.extract(r"(-?\d+(\.\d+)?)", expand=True)[0]
    return pd.to_numeric(s, errors="coerce")

def to_date_series(s):
    if s is None:
        return pd.Series(dtype="datetime64[ns]")
    return pd.to_datetime(s, errors="coerce", dayfirst=True)

def categorize_horizon(days):
    if pd.isna(days):
        return None
    if days <= 360:
        return "Curto (até 360d)"
    if days <= 1080:
        return "Médio (361 a 1080d)"
    return "Longo (acima de 1080d)"

def classify_indexer_bancario(raw):
    if raw is None or pd.isna(raw):
        return None
    t = str(raw).upper()
    if "IPCA" in t:
        return "IPCA"
    if "CDI" in t or "PÓS" in t or "POS" in t:
        return "Pós (CDI)"
    if "PRÉ" in t or "PRE" in t or "FIXA" in t:
        return "Pré"
    return None

def parse_rate_value(x):
    if x is None or pd.isna(x):
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).upper().replace("%", "").replace(" ", "")
    m = re.search(r"(-?\d[\d\.,]*)", s)
    if not m:
        return None
    num = m.group(1)
    if "." in num and "," in num:
        num = num.replace(".", "").replace(",", ".")
    elif "," in num:
        num = num.replace(",", ".")
    try:
        return float(num)
    except:
        return None

def format_rate_for_display(rate_num, indexador):
    if rate_num is None or pd.isna(rate_num):
        return ""
    val = float(rate_num)

    if indexador == "Pós (CDI)":
        val = val * 100 if val <= 2 else val
        return f"{val:,.2f}%".replace(",", "X").replace(".", ",").replace("X", ".")

    if val <= 1.5:
        val = val * 100
    return f"{val:.2f}%".replace(".", ",")

def format_currency_brl(value):
    if value is None or pd.isna(value):
        return ""
    try:
        v = float(value)
    except:
        return ""
    return f"R$ {int(round(v)):,.0f}".replace(",", ".")

def format_date_br(dt):
    if dt is None or pd.isna(dt):
        return ""
    return pd.to_datetime(dt).strftime("%d/%m/%Y")

def copy_button(text: str, label: str = "Copiar"):
    safe = text.replace("\\", "\\\\").replace("`", "\\`").replace("${", "\\${")
    html = f"""
    <button onclick="navigator.clipboard.writeText(`{safe}`)"
    style="cursor:pointer;padding:8px 12px;border-radius:10px;border:1px solid #ddd;background:white;">
    📋 {label}
    </button>
    """
    components.html(html, height=46)

# =============================
# Z-API (send + metadata)
# =============================
def zapi_send_text(
    instance_id: str,
    instance_token: str,
    client_token: str,
    phone_or_group: str,
    message: str,
    delay_message: int = 0,
    mentioned: list[str] | None = None,
):
    url = f"https://api.z-api.io/instances/{instance_id}/token/{instance_token}/send-text"
    headers = {
        "Client-Token": client_token,
        "Content-Type": "application/json",
    }
    payload = {"phone": phone_or_group, "message": message}

    if mentioned:
        payload["mentioned"] = mentioned

    if delay_message and 1 <= int(delay_message) <= 15:
        payload["delayMessage"] = int(delay_message)

    r = requests.post(url, headers=headers, json=payload, timeout=60)
    r.raise_for_status()
    return r.json()

def zapi_group_metadata(instance_id: str, instance_token: str, client_token: str, group_id: str):
    url = f"https://api.z-api.io/instances/{instance_id}/token/{instance_token}/group-metadata"
    headers = {"Client-Token": client_token}
    params = {"phone": group_id}
    r = requests.get(url, headers=headers, params=params, timeout=60)
    r.raise_for_status()
    return r.json()

def extract_participants_phones(metadata_json: dict) -> list[str]:
    participants = metadata_json.get("participants", []) or metadata_json.get("group", {}).get("participants", [])
    phones = []
    for p in participants:
        if isinstance(p, dict):
            ph = p.get("phone")
            if ph:
                phones.append(str(ph))
        elif isinstance(p, str):
            phones.append(p)
    phones = [re.sub(r"\D", "", x) for x in phones if x]
    phones = list(dict.fromkeys([x for x in phones if x]))
    return phones

@st.cache_data(show_spinner=False, ttl=600)
def cached_group_participants(instance_id: str, instance_token: str, client_token: str, group_id: str):
    meta = zapi_group_metadata(instance_id, instance_token, client_token, group_id)
    return extract_participants_phones(meta)

# =============================
# Excel reader
# =============================
@st.cache_data(show_spinner=False)
def read_sheet_fast(file_bytes: bytes, sheet_name: str, header_row: int) -> pd.DataFrame:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f'Aba "{sheet_name}" não encontrada. Abas: {wb.sheetnames}')
    ws = wb[sheet_name]

    header = [normalize_colname(c.value) for c in ws[header_row]]

    data = []
    empty_streak = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row is None or all(x is None or str(x).strip() == "" for x in row):
            empty_streak += 1
            if empty_streak >= 20:
                break
            continue
        empty_streak = 0
        data.append(row)

    df = pd.DataFrame(data, columns=header)
    return df.dropna(axis=1, how="all")

# =============================
# Upload
# =============================
uploaded = st.file_uploader("Envie a planilha (.xlsx ou .xlsm)", type=["xlsx", "xlsm"])

# =============================
# Sidebar config + secrets
# =============================
with st.sidebar:
    st.header("Configurações")

    top_n_banc = st.number_input(
        "Top N por bloco (Crédito bancário)",
        min_value=1,
        max_value=20,
        value=5,
        step=1
    )
    mostrar_apenas_blocos_com_ativos = st.checkbox("Mostrar apenas blocos com ativos", value=True)

    st.divider()
    st.subheader("Envio WhatsApp (Z-API) via secrets")

    secrets_ok = True
    try:
        z_instance_id = st.secrets["zapi"]["instance_id"]
        z_instance_token = st.secrets["zapi"]["instance_token"]
        z_client_token = st.secrets["zapi"]["client_token"]
        groups_dict = dict(st.secrets["groups"])  # {nome: groupId}
    except Exception:
        secrets_ok = False
        z_instance_id = z_instance_token = z_client_token = ""
        groups_dict = {}

    if secrets_ok and groups_dict:
        st.success("Credenciais carregadas via secrets ✅")
        st.caption("Grupos configurados (nomes):")
        for nome_grupo in groups_dict.keys():
            st.write(f"• {nome_grupo.replace('_', ' ').title()}")
    else:
        st.error("Secrets não configurados corretamente ❌")
        st.caption('Esperado no secrets: [zapi] e [groups].')

    delay_between = st.number_input("Pausa entre mensagens (seg)", min_value=0.0, value=2.0, step=0.5)
    api_delay_message = st.number_input("delayMessage (Z-API 0-15)", min_value=0, max_value=15, value=0, step=1)

    st.divider()
    st.subheader("Menções (@)")

    mention_all_enabled = st.checkbox("Ativar menção de participantes", value=False)

    max_mentions = st.number_input(
        "Máx. menções por mensagem",
        min_value=0,
        max_value=2000,
        value=50,
        step=10
    )

    mention_groups = []
    if mention_all_enabled and groups_dict:
        mention_groups = st.multiselect(
            "Mencionar todos apenas nestes grupos:",
            options=list(groups_dict.keys()),
            default=[],
        )
        st.caption("Dica: selecione só grupos internos. Em grupos de clientes, use poucas menções.")

if not uploaded:
    st.info("Envie o arquivo para começar.")
    st.stop()

file_bytes = uploaded.getvalue()

tab_banc, tab_pub = st.tabs(["Crédito bancário", "Títulos públicos (NTN-B)"])

# =========================================================
# TAB 1: Crédito bancário
# =========================================================
with tab_banc:
    st.subheader("Crédito bancário")

    df = read_sheet_fast(file_bytes, SHEET_BANCARIO, header_row=HEADER_BANCARIO)

    col_emissor = find_col(df, ["Emissor"])
    col_produto = find_col(df, ["Produto"])
    col_indexador = find_col(df, ["Indexador"])
    col_taxa = find_col(df, ["Tx. Portal", "Taxa Portal"])
    col_prazo = find_col(df, ["Prazo"])
    col_venc = find_col(df, ["Vencimento"])
    col_min = find_col(df, ["Aplicação", "Aplicacao", "mínima", "minima"])

    missing = []
    if not col_emissor: missing.append("Emissor")
    if not col_produto: missing.append("Produto")
    if not col_indexador: missing.append("Indexador")
    if not col_taxa: missing.append("Tx. Portal/Taxa Portal")
    if not col_prazo and not col_venc: missing.append("Prazo ou Vencimento")
    if not col_min: missing.append("Aplicação mínima")
    if not col_venc: missing.append("Vencimento")

    if missing:
        st.error("Colunas obrigatórias não encontradas: " + ", ".join(missing))
        st.write("Colunas detectadas:", list(df.columns))
        st.stop()

    df["indexador_pad"] = df[col_indexador].apply(classify_indexer_bancario)

    if col_prazo:
        df["prazo_dias"] = to_numeric_series(df[col_prazo])
    else:
        venc_dt = to_date_series(df[col_venc])
        df["prazo_dias"] = (venc_dt - pd.Timestamp(date.today())).dt.days

    df["horizonte"] = df["prazo_dias"].apply(categorize_horizon)

    df["taxa_num"] = df[col_taxa].apply(parse_rate_value)
    df["taxa_fmt"] = df.apply(lambda r: format_rate_for_display(r["taxa_num"], r["indexador_pad"]), axis=1)

    df["aplic_min_num"] = to_numeric_series(df[col_min])
    df["aplic_min_fmt"] = df["aplic_min_num"].apply(format_currency_brl)

    df["_venc_dt"] = to_date_series(df[col_venc])
    df["venc_fmt"] = df["_venc_dt"].apply(format_date_br)

    df = df[df["indexador_pad"].notna() & df["horizonte"].notna() & df["taxa_num"].notna()].copy()

    st.markdown("##### Base tratada (preview)")
    st.dataframe(
        df[[col_emissor, col_produto, col_indexador, "taxa_fmt", "aplic_min_fmt", "venc_fmt", "horizonte"]].head(80),
        use_container_width=True,
        height=340,
    )

    indexers = ["Pós (CDI)", "Pré", "IPCA"]
    horizons = ["Curto (até 360d)", "Médio (361 a 1080d)", "Longo (acima de 1080d)"]

    def top_n_block(dff, idx, horizon, n):
        sub = dff[(dff["indexador_pad"] == idx) & (dff["horizonte"] == horizon)]
        return sub.sort_values("taxa_num", ascending=False).head(int(n))

    st.divider()
    st.subheader("Top do dia (Top N por indexador e horizonte)")

    tabs = st.tabs(indexers)
    for i, idx in enumerate(indexers):
        with tabs[i]:
            cols = st.columns(3)
            for j, hz in enumerate(horizons):
                with cols[j]:
                    st.markdown(f"### {hz}")
                    b = top_n_block(df, idx, hz, top_n_banc)
                    if b.empty:
                        st.info("Sem ativos")
                    else:
                        st.dataframe(
                            b[[col_emissor, col_produto, col_indexador, "taxa_fmt", "aplic_min_fmt", "venc_fmt"]],
                            use_container_width=True,
                            height=260,
                        )

    st.divider()
    st.subheader("Mensagens prontas para WhatsApp (Crédito bancário)")

    TOP_WA = 5

    def format_card_banc(row, prefixo_taxa=""):
        emissor = str(row.get(col_emissor, "")).strip()
        produto = str(row.get(col_produto, "")).strip()
        taxa = str(row.get("taxa_fmt", "")).strip()
        venc = str(row.get("venc_fmt", "")).strip()
        amin = str(row.get("aplic_min_fmt", "")).strip()

        titulo = f"{produto} {emissor}".strip()
        taxa_exibicao = f"{prefixo_taxa}{taxa}" if taxa else ""

        return (
            f"🏦*{titulo}*\n"
            f"⏰ Vencimento: {venc}\n"
            f"📈 Taxa: {taxa_exibicao}\n"
            f"💰mínimo: {amin}\n"
        )

    def build_message_bancario(indexador_label, titulo_indexador, prefixo_taxa=""):
        data_envio = datetime.now().strftime("%d/%m/%Y")
        msg = (
            "*Destaques de ativos Bancários*\n"
            f"🚨*TAXAS DE HOJE ({data_envio})*\n\n"
            f"📍*{titulo_indexador}*\n\n"
        )

        for hz_label, hz_title in [
            ("Curto (até 360d)", "Curto Prazo (até 360d)"),
            ("Médio (361 a 1080d)", "Médio Prazo (361 a 1080d)"),
            ("Longo (acima de 1080d)", "Longo Prazo (acima de 1080d)"),
        ]:
            sub = top_n_block(df, indexador_label, hz_label, TOP_WA)

            if sub.empty and mostrar_apenas_blocos_com_ativos:
                continue

            msg += f"*{hz_title}*\n\n"
            if sub.empty:
                msg += "- (sem ativos hoje)\n\n"
            else:
                for _, r in sub.iterrows():
                    msg += format_card_banc(r, prefixo_taxa=prefixo_taxa) + "\n"
                msg += "\n"

        return msg

    msg_pos = build_message_bancario("Pós (CDI)", "PÓS-FIXADOS")
    msg_pre = build_message_bancario("Pré", "PRÉ-FIXADOS")
    msg_ipca = build_message_bancario("IPCA", "IPCA", prefixo_taxa="IPCA+ ")

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown("### Pós-fixados")
        st.text_area("Mensagem Pós", value=msg_pos, height=560)
        copy_button(msg_pos, "Copiar Pós")
    with c2:
        st.markdown("### Pré-fixados")
        st.text_area("Mensagem Pré", value=msg_pre, height=560)
        copy_button(msg_pre, "Copiar Pré")
    with c3:
        st.markdown("### IPCA")
        st.text_area("Mensagem IPCA", value=msg_ipca, height=560)
        copy_button(msg_ipca, "Copiar IPCA")

# =========================================================
# TAB 2: Títulos públicos (somente NTN-B, listar todas)
# =========================================================
with tab_pub:
    st.subheader("Títulos públicos (Cliente | somente NTN-B | listar todas)")

    dfp = read_sheet_fast(file_bytes, SHEET_PUBLICOS, header_row=HEADER_PUBLICOS)

    col_titulo = find_col(dfp, ["Título"])
    col_venc = find_col(dfp, ["Vencimento"])
    col_taxa = find_col(dfp, ["Taxa do portal às 10h", "Taxa do portal às 10h ¹", "Taxa do portal"])

    missing_pub = []
    if not col_titulo: missing_pub.append("Título")
    if not col_venc: missing_pub.append("Vencimento")
    if not col_taxa: missing_pub.append("Taxa do portal às 10h")

    if missing_pub:
        st.error("Títulos públicos: colunas obrigatórias não encontradas: " + ", ".join(missing_pub))
        st.write("Colunas detectadas:", list(dfp.columns))
        st.stop()

    dfp = dfp[dfp[col_titulo].astype(str).str.upper().str.contains("NTN-B")].copy()

    dfp["_venc_dt"] = to_date_series(dfp[col_venc])
    dfp["venc_fmt"] = dfp["_venc_dt"].apply(format_date_br)

    dfp["prazo_dias"] = (dfp["_venc_dt"] - pd.Timestamp(date.today())).dt.days
    dfp["horizonte"] = dfp["prazo_dias"].apply(categorize_horizon)

    dfp["taxa_num"] = dfp[col_taxa].apply(parse_rate_value)
    dfp["taxa_fmt"] = dfp["taxa_num"].apply(lambda x: format_rate_for_display(x, indexador="IPCA"))

    dfp = dfp[dfp["horizonte"].notna() & dfp["taxa_num"].notna()].copy()

    st.markdown("##### Preview (somente NTN-B)")
    st.dataframe(
        dfp[[col_titulo, "venc_fmt", "taxa_fmt", "horizonte", "prazo_dias"]]
        .sort_values("prazo_dias")
        .head(150),
        use_container_width=True,
        height=340,
    )

    def pub_block_all(hz):
        sub = dfp[dfp["horizonte"] == hz].copy()
        return sub.sort_values("prazo_dias")

    def format_card_pub(row):
        titulo = str(row.get(col_titulo, "")).strip()
        venc = str(row.get("venc_fmt", "")).strip()
        taxa = str(row.get("taxa_fmt", "")).strip()
        return (
            f"🏛️*{titulo}*\n"
            f"⏰ Vencimento: {venc}\n"
            f"📈 Taxa: IPCA+ {taxa}\n"
        )

    def build_message_pub_ntnb_all():
        hoje = datetime.now().strftime("%d/%m/%Y")
        msg = (
            "*Destaques de Títulos Públicos*\n"
            f"🚨*TAXAS DE HOJE ({hoje})*\n\n"
            "📍*TESOURO IPCA+ (NTN-B)*\n\n"
        )

        for hz_label, hz_title in [
            ("Curto (até 360d)", "Curto Prazo (até 360d)"),
            ("Médio (361 a 1080d)", "Médio Prazo (361 a 1080d)"),
            ("Longo (acima de 1080d)", "Longo Prazo (acima de 1080d)"),
        ]:
            sub = pub_block_all(hz_label)

            if sub.empty and mostrar_apenas_blocos_com_ativos:
                continue

            msg += f"*{hz_title}*\n\n"
            if sub.empty:
                msg += "- (sem títulos hoje)\n\n"
            else:
                for _, r in sub.iterrows():
                    msg += format_card_pub(r) + "\n"
                msg += "\n"

        return msg

    msg_pub_ntnb = build_message_pub_ntnb_all()

    st.divider()
    st.subheader("Mensagem pronta para WhatsApp | Tesouro IPCA+ (todas as NTN-B)")
    st.text_area("Mensagem Tesouro IPCA+", value=msg_pub_ntnb, height=560)
    copy_button(msg_pub_ntnb, "Copiar Tesouro IPCA+")

# =========================================================
# SEND SECTION (todos os grupos do secrets)
# - Envia para todos os grupos
# - Só menciona participantes nos grupos selecionados
# =========================================================
st.divider()
st.subheader("Enviar para todos os grupos (1 clique)")

messages_to_send = []
for name in ["msg_pos", "msg_pre", "msg_ipca", "msg_pub_ntnb"]:
    if name in globals():
        messages_to_send.append(globals()[name])

can_send = secrets_ok and bool(groups_dict) and len(messages_to_send) >= 1

info_cols = st.columns([2, 3])
with info_cols[0]:
    st.caption("Resumo")
    st.write(f"Mensagens no pacote: {len(messages_to_send)}")
    st.write(f"Grupos no secrets: {len(groups_dict) if groups_dict else 0}")

with info_cols[1]:
    if mention_all_enabled:
        st.caption("Menções")
        if mention_groups:
            nice = ", ".join([g.replace("_", " ").title() for g in mention_groups])
            st.write(f"Mencionar participantes apenas em: {nice}")
        else:
            st.write("Nenhum grupo selecionado para menção.")
    else:
        st.caption("Menções desativadas.")

col_send1, col_send2 = st.columns([1, 2])

with col_send1:
    if st.button("📤 Enviar mensagens agora", disabled=not can_send):
        all_results = []
        for gname, gid in groups_dict.items():
            mentioned_list = None
            used_mentions = False

            if mention_all_enabled and (gname in (mention_groups or [])):
                try:
                    phones = cached_group_participants(z_instance_id, z_instance_token, z_client_token, gid)
                    if int(max_mentions) > 0:
                        phones = phones[: int(max_mentions)]
                    mentioned_list = phones
                    used_mentions = True
                except Exception as e:
                    mentioned_list = None
                    used_mentions = False
                    st.warning(f"Não consegui carregar participantes do grupo '{gname}'. Enviando sem menção. Erro: {e}")

            group_results = []
            for i, msg in enumerate(messages_to_send, start=1):
                if not msg or not str(msg).strip():
                    continue
                try:
                    res = zapi_send_text(
                        instance_id=z_instance_id,
                        instance_token=z_instance_token,
                        client_token=z_client_token,
                        phone_or_group=gid,
                        message=msg,
                        delay_message=api_delay_message if api_delay_message > 0 else 0,
                        mentioned=mentioned_list,
                    )
                    group_results.append({"ordem": i, "ok": True, "response": res})
                except requests.HTTPError as e:
                    group_results.append({
                        "ordem": i,
                        "ok": False,
                        "status_code": getattr(e.response, "status_code", None),
                        "error_text": getattr(e.response, "text", str(e)),
                    })
                except Exception as e:
                    group_results.append({"ordem": i, "ok": False, "error_text": str(e)})

                if delay_between:
                    time.sleep(float(delay_between))

            all_results.append({
                "grupo": gname,
                "mencionou_todos": used_mentions,
                "qtd_mencoes": 0 if not mentioned_list else len(mentioned_list),
                "resultados": group_results
            })

        st.success(f"Envio concluído. Total de grupos: {len(all_results)}")
        st.json(all_results)

with col_send2:
    st.caption(
        "Dica: se enviar para muitos grupos, aumente a pausa. "
        "Se marcar pessoas, use poucas menções em grupos de clientes."
    )
